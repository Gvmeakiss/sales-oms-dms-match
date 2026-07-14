#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将SQL结果数据整理为符合会计规范的Excel
- 解析订单和发货SQL文件
- 按照会计规范格式整理数据
- 生成符合会计规范的Excel文件
"""

import pandas as pd
import numpy as np
from pathlib import Path
import chardet
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def get_encoding(file_path):
    """检测文件编码"""
    with open(file_path, 'rb') as f:
        return chardet.detect(f.read())['encoding']

def _parse_values_line(line):
    """从 VALUES 行解析出值列表，处理引号与 NULL"""
    line = line.strip()
    if not line.upper().startswith('VALUES'):
        return None
    part = line[6:].strip().strip(';')
    start = part.find('(')
    end = part.rfind(')')
    if start < 0 or end <= start:
        return None
    part = part[start + 1 : end]
    if not part.strip():
        return None
    parts = []
    cur = ''
    in_q = False
    for c in part:
        if c == "'":
            in_q = not in_q
            cur += c
        elif c == ',' and not in_q:
            if cur.strip():
                parts.append(cur.strip())
            cur = ''
        else:
            cur += c
    if cur.strip():
        parts.append(cur.strip())
    out = []
    for p in parts:
        p = p.strip()
        if p.upper() == 'NULL':
            out.append(None)
        else:
            if p.startswith("'") and p.endswith("'"):
                p = p[1:-1]
            out.append(p)
    return out

def parse_order_sql(sql_path):
    """解析订单SQL文件"""
    print(f"\n正在解析订单SQL文件: {sql_path}")
    enc = get_encoding(sql_path)
    with open(sql_path, 'r', encoding=enc, errors='ignore') as f:
        lines = f.readlines()
    
    # 从第一行获取列名
    first_line = lines[0] if lines else ""
    if 'INSERT INTO' in first_line.upper():
        # 提取列名
        start = first_line.find('(')
        end = first_line.find(')')
        if start > 0 and end > start:
            cols_str = first_line[start+1:end]
            expected_cols = [c.strip() for c in cols_str.split(',')]
        else:
            # 默认13列格式
            expected_cols = [
                'platform_order_no', 'main_order_no', 'sale_order_no', 'order_type', 'order_status',
                'create_time', 'update_time', 'channel_name', 'item_code', 'line_amount', 
                'pay_amount', 'item_num', 'organization_code'
            ]
    else:
        expected_cols = [
            'platform_order_no', 'main_order_no', 'sale_order_no', 'order_type', 'order_status',
            'create_time', 'update_time', 'channel_name', 'item_code', 'line_amount', 
            'pay_amount', 'item_num', 'organization_code'
        ]
    
    rows = []
    for line in lines:
        vals = _parse_values_line(line)
        if vals is None:
            continue
        if len(vals) >= len(expected_cols):
            rows.append(vals[:len(expected_cols)])
        elif len(vals) >= 11:  # 11列格式
            # 补全到13列
            if len(vals) == 11:
                vals.insert(9, None)  # 插入line_amount位置
                vals.append(None)  # 添加organization_code
            rows.append(vals[:len(expected_cols)])
    
    df = pd.DataFrame(rows)
    if len(df.columns) != len(expected_cols):
        df = df.iloc[:, :len(expected_cols)]
    df.columns = expected_cols
    
    print(f"  解析完成: {len(df):,} 行, {len(df.columns)} 列")
    return df

def parse_delivery_sql(sql_path):
    """解析发货SQL文件"""
    print(f"\n正在解析发货SQL文件: {sql_path}")
    enc = get_encoding(sql_path)
    with open(sql_path, 'r', encoding=enc, errors='ignore') as f:
        lines = f.readlines()
    
    # 从第一行获取列名
    first_line = lines[0] if lines else ""
    if 'INSERT INTO' in first_line.upper():
        start = first_line.find('(')
        end = first_line.find(')')
        if start > 0 and end > start:
            cols_str = first_line[start+1:end]
            expected_cols = [c.strip() for c in cols_str.split(',')]
        else:
            expected_cols = ['business_type', '订单号', 'external_order_no', '业务时间', '料号', '名称', '已发货数量']
    else:
        expected_cols = ['business_type', '订单号', 'external_order_no', '业务时间', '料号', '名称', '已发货数量']
    
    rows = []
    for line in lines:
        vals = _parse_values_line(line)
        if vals is None:
            continue
        if len(vals) >= 7:
            rows.append(vals[:7])
    
    df = pd.DataFrame(rows)
    if len(df.columns) != len(expected_cols):
        df = df.iloc[:, :len(expected_cols)]
    df.columns = expected_cols
    
    print(f"  解析完成: {len(df):,} 行, {len(df.columns)} 列")
    return df

def format_accounting_excel(df_order, df_delivery, output_path):
    """将数据整理为符合会计规范的Excel格式"""
    print(f"\n正在生成符合会计规范的Excel文件...")
    
    # 创建Excel写入器
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # ========== Sheet 1: 订单明细账 ==========
        print("  生成订单明细账...")
        df_order_formatted = df_order.copy()
        
        # 日期格式化
        if 'create_time' in df_order_formatted.columns:
            df_order_formatted['业务日期'] = pd.to_datetime(
                df_order_formatted['create_time'], errors='coerce'
            ).dt.strftime('%Y-%m-%d')
        else:
            df_order_formatted['业务日期'] = ''
        
        # 金额格式化（保留两位小数）
        amount_cols = ['line_amount', 'pay_amount']
        for col in amount_cols:
            if col in df_order_formatted.columns:
                df_order_formatted[col] = pd.to_numeric(
                    df_order_formatted[col], errors='coerce'
                ).fillna(0).round(2)
        
        # 数量格式化
        if 'item_num' in df_order_formatted.columns:
            df_order_formatted['item_num'] = pd.to_numeric(
                df_order_formatted['item_num'], errors='coerce'
            ).fillna(0)
        
        # 重新排列列顺序（符合会计规范）
        order_cols = [
            '业务日期', 'platform_order_no', 'sale_order_no', 'channel_name',
            'item_code', 'item_num', 'line_amount', 'pay_amount', 
            'order_type', 'order_status', 'create_time', 'update_time'
        ]
        # 只保留存在的列
        order_cols = [col for col in order_cols if col in df_order_formatted.columns]
        df_order_formatted = df_order_formatted[order_cols]
        
        # 重命名列名为中文（符合会计规范）
        rename_dict = {
            '业务日期': '业务日期',
            'platform_order_no': '平台订单号',
            'sale_order_no': '销售订单号',
            'channel_name': '渠道名称',
            'item_code': '商品代码',
            'item_num': '商品数量',
            'line_amount': '行金额',
            'pay_amount': '支付金额',
            'order_type': '订单类型',
            'order_status': '订单状态',
            'create_time': '创建时间',
            'update_time': '更新时间'
        }
        df_order_formatted.rename(columns=rename_dict, inplace=True)
        
        df_order_formatted.to_excel(
            writer, sheet_name='订单明细账', index=False, na_rep=''
        )
        
        # ========== Sheet 2: 发货明细账 ==========
        print("  生成发货明细账...")
        df_delivery_formatted = df_delivery.copy()
        
        # 日期格式化
        if '业务时间' in df_delivery_formatted.columns:
            df_delivery_formatted['业务日期'] = pd.to_datetime(
                df_delivery_formatted['业务时间'], errors='coerce'
            ).dt.strftime('%Y-%m-%d')
        else:
            df_delivery_formatted['业务日期'] = ''
        
        # 数量格式化
        if '已发货数量' in df_delivery_formatted.columns:
            df_delivery_formatted['已发货数量'] = pd.to_numeric(
                df_delivery_formatted['已发货数量'], errors='coerce'
            ).fillna(0)
        
        # 重新排列列顺序
        delivery_cols = [
            '业务日期', 'business_type', '订单号', 'external_order_no',
            '料号', '名称', '已发货数量', '业务时间'
        ]
        delivery_cols = [col for col in delivery_cols if col in df_delivery_formatted.columns]
        df_delivery_formatted = df_delivery_formatted[delivery_cols]
        
        # 重命名列名为中文
        rename_dict_delivery = {
            '业务日期': '业务日期',
            'business_type': '业务类型',
            '订单号': '订单号',
            'external_order_no': '外部订单号',
            '料号': '料号',
            '名称': '商品名称',
            '已发货数量': '已发货数量',
            '业务时间': '业务时间'
        }
        df_delivery_formatted.rename(columns=rename_dict_delivery, inplace=True)
        
        df_delivery_formatted.to_excel(
            writer, sheet_name='发货明细账', index=False, na_rep=''
        )
        
        # ========== Sheet 3: 汇总统计 ==========
        print("  生成汇总统计...")
        summary_data = []
        
        # 订单汇总
        if 'pay_amount' in df_order.columns:
            total_order_amount = pd.to_numeric(df_order['pay_amount'], errors='coerce').sum()
            summary_data.append(['订单总金额', f'{total_order_amount:,.2f}', '元'])
        
        if 'item_num' in df_order.columns:
            total_order_qty = pd.to_numeric(df_order['item_num'], errors='coerce').sum()
            summary_data.append(['订单总数量', f'{total_order_qty:,.0f}', '件'])
        
        summary_data.append(['订单记录数', f'{len(df_order):,}', '条'])
        
        # 发货汇总
        if '已发货数量' in df_delivery.columns:
            total_delivery_qty = pd.to_numeric(df_delivery['已发货数量'], errors='coerce').sum()
            summary_data.append(['发货总数量', f'{total_delivery_qty:,.0f}', '件'])
        
        summary_data.append(['发货记录数', f'{len(df_delivery):,}', '条'])
        
        df_summary = pd.DataFrame(summary_data, columns=['项目', '数值', '单位'])
        df_summary.to_excel(writer, sheet_name='汇总统计', index=False)
    
    # 格式化Excel
    print("  正在格式化Excel...")
    wb = load_workbook(output_path)
    
    # 定义样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    
    # 格式化每个工作表
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 设置表头样式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # 设置列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # 设置数据格式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                # 金额列右对齐
                if '金额' in str(ws[cell.row][0].value) or '金额' in str(cell.value):
                    cell.alignment = right_alignment
                # 数量列右对齐
                elif '数量' in str(ws[cell.row][0].value) or '数量' in str(cell.value):
                    cell.alignment = right_alignment
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 冻结首行
        ws.freeze_panes = 'A2'
    
    wb.save(output_path)
    print(f"\n✓ Excel文件已生成: {output_path}")

def main():
    """主函数"""
    print("=" * 60)
    print("SQL数据转会计规范Excel工具")
    print("=" * 60)
    
    # 文件路径
    base_dir = Path(r"D:\0.DA\MiaokeDA\Sales To B\OMS25年1-12月订单及发货数据")
    order_sql = base_dir / "25年10-12月订单数据.sql"
    delivery_sql = base_dir / "25年10-12月发货数据.sql"
    output_excel = base_dir / "25年10-12月订单及发货数据_会计规范格式.xlsx"
    
    # 检查文件是否存在
    if not order_sql.exists():
        print(f"错误: 订单SQL文件不存在: {order_sql}")
        return
    
    if not delivery_sql.exists():
        print(f"错误: 发货SQL文件不存在: {delivery_sql}")
        return
    
    try:
        # 解析SQL文件
        df_order = parse_order_sql(str(order_sql))
        df_delivery = parse_delivery_sql(str(delivery_sql))
        
        # 生成Excel
        format_accounting_excel(df_order, df_delivery, str(output_excel))
        
        print("\n" + "=" * 60)
        print("处理完成！")
        print("=" * 60)
        
    except Exception as e:
        print(f"\n错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
